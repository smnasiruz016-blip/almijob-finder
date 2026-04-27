from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def classify_category(website: str, category: str) -> tuple[str, bool, bool, bool]:
    website_name = (website or "").strip().lower()
    raw = (category or "").strip().lower()

    is_aggregator = "aggregator" in raw or website_name in {"indeed", "jooble", "ziprecruiter"}
    has_api = website_name in {"jooble", "adzuna", "remoteok", "remote ok", "remotive"}

    if "government" in raw:
      return ("government", has_api, is_aggregator, False)

    if "employer" in raw or "company" in raw:
      return ("employer", has_api, is_aggregator, True)

    if "professional" in raw or "linkedin" in website_name or "xing" in website_name:
      return ("professional", has_api, is_aggregator, False)

    if is_aggregator:
      return ("aggregator", has_api, True, False)

    if "local" in raw or "regional" in raw:
      return ("local", has_api, is_aggregator, False)

    return ("general", has_api, is_aggregator, False)


def source_priority(website: str, normalized_category: str) -> int:
    website_name = (website or "").strip().lower()

    if website_name == "linkedin jobs":
      return 10
    if website_name == "jooble":
      return 20
    if website_name == "indeed":
      return 30
    if normalized_category == "local":
      return 40
    if normalized_category == "professional":
      return 50
    if normalized_category == "aggregator":
      return 60
    return 80


def export_sources(input_path: Path, output_path: Path) -> int:
    workbook = load_workbook(input_path, data_only=True)
    worksheet = workbook["Job Websites"]

    rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    records = []

    for region, country, website, url, category, notes in rows:
        if not country or not website or not url:
            continue

        normalized_category, has_api, is_aggregator, is_employer_board = classify_category(str(website), str(category or ""))
        records.append(
            {
                "region": str(region or "").strip() or "Global",
                "country": str(country).strip(),
                "website": str(website).strip(),
                "url": str(url).strip(),
                "category": normalized_category,
                "notes": str(notes or "").strip(),
                "sourcePriority": source_priority(str(website), normalized_category),
                "hasApi": has_api,
                "isAggregator": is_aggregator,
                "isEmployerBoard": is_employer_board,
                "isTrusted": True,
                "active": True,
            }
        )

    records.extend(
        [
            {
                "region": "Global",
                "country": "Worldwide",
                "website": "LinkedIn Jobs",
                "url": "https://www.linkedin.com/jobs/",
                "category": "professional",
                "notes": "Professional roles and international employers across many markets.",
                "sourcePriority": 10,
                "hasApi": False,
                "isAggregator": False,
                "isEmployerBoard": False,
                "isTrusted": True,
                "active": True,
            },
            {
                "region": "Global",
                "country": "Worldwide",
                "website": "Jooble",
                "url": "https://jooble.org/",
                "category": "aggregator",
                "notes": "Broad worldwide search coverage across many countries.",
                "sourcePriority": 20,
                "hasApi": True,
                "isAggregator": True,
                "isEmployerBoard": False,
                "isTrusted": True,
                "active": True,
            },
            {
                "region": "Global",
                "country": "Worldwide",
                "website": "Indeed",
                "url": "https://www.indeed.com/",
                "category": "aggregator",
                "notes": "Strong general job coverage and useful for broader searches.",
                "sourcePriority": 30,
                "hasApi": False,
                "isAggregator": True,
                "isEmployerBoard": False,
                "isTrusted": True,
                "active": True,
            },
        ]
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
    return len(records)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python export_job_source_directory.py <input.xlsx> <output.json>")

    count = export_sources(Path(sys.argv[1]), Path(sys.argv[2]))
    print(f"Exported {count} job source rows.")
